import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# This Whole code help us to make our Timetable looking More Visually Good And separate time table for each Branch or division 

def format_timetable(input_file="generated_timetable.xlsx", output_file="Formatted_Timetable.xlsx"):
    """
    Converts raw timetable data into a formatted timetable with:
    - Days as rows
    - Time slots as columns
    - Separate sheets for each division
    - Color-coded cells based on division
    - Merged cells for multi-hour courses (e.g., labs)
    
    Args:
        input_file: Path to the raw Excel file
        output_file: Path for the output Excel file
    """
    print(f"Reading raw timetable from {input_file}...")
    try:
        raw_df = pd.read_excel(input_file)
    except FileNotFoundError:
        print(f"Error: {input_file} not found.")
        return
    except Exception as e:
        print(f"Error reading file: {e}")
        return
    
    # Color map for divisions
    COLOR_MAP = {
        "Div 1-2": "FFB6C1",      # Light Pink
        "Div 3-4": "ADD8E6",      # Light Blue
        "AI 2Y": "90EE90",        # Light Green
        "CE 2Y": "FFFFE0",        # Light Yellow
        "CH 2Y": "D8BFD8",        # Thistle
        "CSE 2Y": "FFD700",       # Gold
        "Common": "F0F0F0"        # Light Grey
    }
    
    # Define time slots and days - match exactly with first script
    TIME_SLOTS = ["9:00 AM", "10:00 AM", "11:00 AM", "12:00 PM", "2:00 PM", "3:00 PM", "4:00 PM", "5:00 PM"]
    DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
    
    # Map abbreviated day names to full names for display
    DAY_MAP = {
        "Mon": "Monday",
        "Tue": "Tuesday",
        "Wed": "Wednesday",
        "Thu": "Thursday",
        "Fri": "Friday"
    }
    
    # Debug: Print unique days and times in the data
    print("Unique days in data:", raw_df["Day"].unique())
    print("Unique times in data:", raw_df["Time"].unique())
    
    # Create a workbook
    wb = Workbook()
    
    # Get unique divisions
    divisions = sorted(raw_df["Division"].unique())
    
    # Create a sheet for each division
    for i, division in enumerate(divisions):
        if i == 0:
            ws = wb.active
            ws.title = division
        else:
            ws = wb.create_sheet(title=division)
        
        print(f"Processing division: {division}")
        
        # Filter data for this division
        div_data = raw_df[raw_df["Division"] == division]
        
        # Set up table structure
        ws.cell(row=1, column=1).value = division
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Add time slots as columns
        for col_idx, time in enumerate(TIME_SLOTS, start=2):
            ws.cell(row=1, column=col_idx).value = time
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")
        
        # Add days as rows (using the abbreviated format that matches raw data)
        for row_idx, day in enumerate(DAYS, start=2):
            ws.cell(row=row_idx, column=1).value = DAY_MAP[day]  # Display full day name
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="center")
        
        # Identify multi-hour courses
        multi_hour_courses = {}
        for _, row in div_data.iterrows():
            course = row["Course"]
            # Extract base course name from "(3hr start)" or "(cont)" or "(end)" formats
            base_course = course.split(" (")[0] if " (" in course else course
            
            if "(3hr start)" in course or "(4hr start)" in course or "(2hr start)" in course:
                day = row["Day"]
                time = row["Time"] 
                multi_hour_courses[(day, base_course)] = {"start_time": time, "duration": 0}
            
            if "(cont)" in course or "(end)" in course:
                day = row["Day"]
                if (day, base_course) in multi_hour_courses:
                    multi_hour_courses[(day, base_course)]["duration"] += 1
        
        # For each day and time, find the corresponding course
        for day in DAYS:
            day_data = div_data[div_data["Day"] == day]
            
            for time_idx, time in enumerate(TIME_SLOTS, start=2):
                time_data = day_data[day_data["Time"] == time]
                
                if not time_data.empty:
                    row_idx = DAYS.index(day) + 2
                    course_row = time_data.iloc[0]
                    
                    course = course_row["Course"]
                    base_course = course.split(" (")[0] if " (" in course else course
                    
                    # Skip if this is a continuation cell that will be merged
                    if "(cont)" in course or "(end)" in course:
                        continue
                    
                    # Create content for the cell
                    teacher = course_row["Teacher"]
                    room = course_row["Room"]
                    course_info = f"{base_course}\n{teacher}\n{room}"
                    
                    cell = ws.cell(row=row_idx, column=time_idx)
                    cell.value = course_info
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    
                    # Apply color based on division
                    fill_color = COLOR_MAP.get(division, "FFFFFF")
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    # Add border
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    # Handle multi-hour courses by merging cells
                    if "(hr start)" in course:
                        # Determine how many cells to merge
                        if "(3hr start)" in course:
                            span = 3
                        elif "(4hr start)" in course:
                            span = 4
                        elif "(2hr start)" in course:
                            span = 2
                        else:
                            span = 1
                        
                        if span > 1 and time_idx + span - 1 <= len(TIME_SLOTS) + 1:
                            try:
                                ws.merge_cells(start_row=row_idx, start_column=time_idx,
                                              end_row=row_idx, end_column=time_idx + span - 1)
                            except Exception as e:
                                print(f"Could not merge cells for {course} at {day} {time}: {e}")
        
        # Adjust column widths and row heights
        ws.column_dimensions[get_column_letter(1)].width = 15
        for col_idx in range(2, len(TIME_SLOTS) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        for row_idx in range(1, len(DAYS) + 2):
            ws.row_dimensions[row_idx].height = 80
    
    # Create Master Timetable and Room View
    create_master_timetable(wb, raw_df, DAYS, TIME_SLOTS, COLOR_MAP, DAY_MAP)
    create_room_view(wb, raw_df, DAYS, TIME_SLOTS, COLOR_MAP, DAY_MAP)
    
    # Save the workbook
    try:
        wb.save(output_file)
        print(f"Formatted timetable saved as {output_file}")
        print("The timetable includes:")
        print("- Individual sheets for each division")
        print("- A master timetable showing all divisions")
        print("- A room-based view")
        print("- Color-coded cells and merged cells for multi-hour courses")
    except Exception as e:
        print(f"Error saving file: {e}")

def create_master_timetable(wb, raw_df, DAYS, TIME_SLOTS, COLOR_MAP, DAY_MAP):
    """Creates a master timetable sheet with all divisions"""
    ws = wb.create_sheet(title="Master Timetable")
    
    ws.cell(row=1, column=1).value = "Master Timetable"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    for col_idx, time in enumerate(TIME_SLOTS, start=2):
        ws.cell(row=1, column=col_idx).value = time
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")
    
    row_offset = 2
    for day in DAYS:
        for room in sorted(raw_df["Room"].unique()):
            # Display full day name with room
            ws.cell(row=row_offset, column=1).value = f"{DAY_MAP[day]} - {room}"
            ws.cell(row=row_offset, column=1).font = Font(bold=True)
            ws.cell(row=row_offset, column=1).alignment = Alignment(vertical="center")
            
            day_room_data = raw_df[(raw_df["Day"] == day) & (raw_df["Room"] == room)]
            
            # Process each time slot
            for time_idx, time in enumerate(TIME_SLOTS, start=2):
                time_data = day_room_data[day_room_data["Time"] == time]
                
                if not time_data.empty:
                    # Skip continuation cells that should be merged
                    course = time_data.iloc[0]["Course"]
                    if "(cont)" in course or "(end)" in course:
                        continue
                    
                    base_course = course.split(" (")[0] if " (" in course else course
                    division = time_data.iloc[0]["Division"]
                    teacher = time_data.iloc[0]["Teacher"]
                    
                    cell = ws.cell(row=row_offset, column=time_idx)
                    cell.value = f"{base_course}\n{division}\n{teacher}"
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    
                    fill_color = COLOR_MAP.get(division, "FFFFFF")
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    # Handle multi-hour courses by merging cells
                    if "(hr start)" in course:
                        # Determine how many cells to merge
                        if "(3hr start)" in course:
                            span = 3
                        elif "(4hr start)" in course:
                            span = 4
                        elif "(2hr start)" in course:
                            span = 2
                        else:
                            span = 1
                        
                        if span > 1 and time_idx + span - 1 <= len(TIME_SLOTS) + 1:
                            try:
                                ws.merge_cells(start_row=row_offset, start_column=time_idx,
                                              end_row=row_offset, end_column=time_idx + span - 1)
                            except Exception as e:
                                print(f"Could not merge cells for {course} at {day} {time}: {e}")
            
            row_offset += 1
    
    # Adjust column widths and row heights
    ws.column_dimensions[get_column_letter(1)].width = 25
    for col_idx in range(2, len(TIME_SLOTS) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    for row_idx in range(1, row_offset):
        ws.row_dimensions[row_idx].height = 50

def create_room_view(wb, raw_df, DAYS, TIME_SLOTS, COLOR_MAP, DAY_MAP):
    """Creates a separate sheet for room-based view"""
    ws = wb.create_sheet(title="Room View")
    
    ws.cell(row=1, column=1).value = "Room Schedule"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    for col_idx, time in enumerate(TIME_SLOTS, start=2):
        ws.cell(row=1, column=col_idx).value = time
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")
    
    rooms = sorted(raw_df["Room"].unique())
    row_idx = 2
    
    for room in rooms:
        for day in DAYS:
            # Display full day name with room
            ws.cell(row=row_idx, column=1).value = f"{room} - {DAY_MAP[day]}"
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="center")
            
            room_day_data = raw_df[(raw_df["Room"] == room) & (raw_df["Day"] == day)]
            
            # Process each time slot
            for time_idx, time in enumerate(TIME_SLOTS, start=2):
                time_data = room_day_data[room_day_data["Time"] == time]
                
                if not time_data.empty:
                    # Skip continuation cells that should be merged
                    course = time_data.iloc[0]["Course"]
                    if "(cont)" in course or "(end)" in course:
                        continue
                    
                    base_course = course.split(" (")[0] if " (" in course else course
                    division = time_data.iloc[0]["Division"]
                    teacher = time_data.iloc[0]["Teacher"]
                    
                    cell = ws.cell(row=row_idx, column=time_idx)
                    cell.value = f"{base_course}\n{division}\n{teacher}"
                    cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                    
                    fill_color = COLOR_MAP.get(division, "FFFFFF")
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    # Handle multi-hour courses by merging cells
                    if "(hr start)" in course:
                        # Determine how many cells to merge
                        if "(3hr start)" in course:
                            span = 3
                        elif "(4hr start)" in course:
                            span = 4
                        elif "(2hr start)" in course:
                            span = 2
                        else:
                            span = 1
                        
                        if span > 1 and time_idx + span - 1 <= len(TIME_SLOTS) + 1:
                            try:
                                ws.merge_cells(start_row=row_idx, start_column=time_idx,
                                              end_row=row_idx, end_column=time_idx + span - 1)
                            except Exception as e:
                                print(f"Could not merge cells for {course} at {day} {time}: {e}")
            
            row_idx += 1
    
    # Adjust column widths and row heights
    ws.column_dimensions[get_column_letter(1)].width = 25
    for col_idx in range(2, len(TIME_SLOTS) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    for row_idx in range(1, len(rooms) * len(DAYS) + 2):
        ws.row_dimensions[row_idx].height = 50

if __name__ == "__main__":
    format_timetable()
